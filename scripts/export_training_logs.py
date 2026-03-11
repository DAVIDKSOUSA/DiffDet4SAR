#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = "output_aircraft_single_class/metrics.json"
DEFAULT_XLSX = "output_aircraft_single_class/metrics.xlsx"
DEFAULT_JSON = "output_aircraft_single_class/metrics_records.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export detectron2 metrics.json (JSONL) to Excel (.xlsx) and JSON array."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help="Path to detectron2 metrics.json (JSONL format).",
    )
    parser.add_argument(
        "--xlsx-output",
        default=DEFAULT_XLSX,
        help="Destination .xlsx path.",
    )
    parser.add_argument(
        "--json-output",
        default=DEFAULT_JSON,
        help="Destination JSON array path.",
    )
    return parser.parse_args()


def load_jsonl(path: Path) -> List[Dict]:
    records: List[Dict] = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        payload = line.strip()
        if not payload:
            continue
        try:
            obj = json.loads(payload)
        except json.JSONDecodeError:
            print(f"Skipping invalid JSON line {line_no}")
            continue
        if isinstance(obj, dict):
            records.append(obj)
    return records


def ordered_columns(records: Sequence[Dict]) -> List[str]:
    keys = set()
    for rec in records:
        keys.update(rec.keys())

    priority = [
        "iteration",
        "eta_seconds",
        "time",
        "data_time",
        "lr",
        "total_loss",
        "loss_ce",
        "loss_bbox",
        "loss_giou",
        "bbox/AP",
        "bbox/AP50",
        "bbox/AP75",
        "bbox/APs",
        "bbox/APm",
        "bbox/APl",
    ]
    columns = [key for key in priority if key in keys]
    columns.extend(sorted(keys - set(columns)))
    return columns


def autosize_columns(worksheet, columns: Sequence[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        width = min(max(len(name) + 2, 12), 50)
        worksheet.column_dimensions[get_column_letter(idx)].width = width


def write_sheet(workbook: Workbook, title: str, records: Sequence[Dict], columns: Sequence[str]) -> None:
    ws = workbook.create_sheet(title=title)
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for rec in records:
        ws.append([rec.get(col) for col in columns])
    ws.freeze_panes = "A2"
    autosize_columns(ws, columns)


def export_to_xlsx(records: Sequence[Dict], output_path: Path) -> None:
    columns = ordered_columns(records)
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "all_metrics", records, columns)

    train_records = [rec for rec in records if "total_loss" in rec]
    if train_records:
        write_sheet(wb, "train_only", train_records, columns)

    eval_records = [rec for rec in records if any(key.startswith("bbox/") for key in rec)]
    if eval_records:
        write_sheet(wb, "eval_only", eval_records, columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_to_json(records: Iterable[Dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(list(records), ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    xlsx_path = Path(args.xlsx_output).expanduser().resolve()
    json_path = Path(args.json_output).expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input metrics file not found: {input_path}")

    records = load_jsonl(input_path)
    if not records:
        raise RuntimeError(f"No valid records found in: {input_path}")

    export_to_xlsx(records, xlsx_path)
    export_to_json(records, json_path)

    print(f"Input records: {len(records)}")
    print(f"Excel exported to: {xlsx_path}")
    print(f"JSON exported to: {json_path}")


if __name__ == "__main__":
    main()
